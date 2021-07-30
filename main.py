from tkinter import *
import train
import face_recognition
from openpyxl import Workbook
from openpyxl import load_workbook
import itertools
from tkinter import filedialog
from PIL import ImageTk, Image
import os

def open_file():
    filename = filedialog.askopenfilename(initialdir="/home/vaibhav/minor_pro/Resources/Faces/val", title="Select An Image", filetypes=(("jpeg files", ".jpg"), ("gif files", ".gif*"), ("png files", "*.png")))
    global d
    d=filename
def open_train():
    train.train()
def open_detect():
    a=face_recognition.recognition(d)
    labeldisplayattendance.config(text=a)


def get_name():
    wb=Workbook()
    wb=load_workbook('Attendance.xlsx')
    ws=wb.active
    col_a=ws['A']
    col_b=ws['B']
    lista=''
    for (cell,cell2) in zip(col_a,col_b):
      lista=f'{lista +str(cell.value)+"                         "+str(cell2.value)}\n'
    labeldisplay.config(text=lista)

root=Tk()
root.geometry("1000x800")
root.title('Attendance System')
root.config(background="white")
labela=Label(root,text="ATTENDANCE SYSTEM",width=100,height=4)
labela.grid(row=0)
buttontrain=Button(root,text="Train model",command=open_train,fg="green")
buttontrain.grid(row=1,pady=20)
labelentry=Label(root,text="Enter file path of student image",width=60,height=3)
labelentry.grid(row=2,pady=20)
buttonopen=Button(root,text="OPEN FILE FOR ATTENDANCE",command=open_file)
buttonopen.grid(row=3,pady=20)
buttondetect=Button(root,text="Detect Attendance",command=open_detect,fg="green")
buttondetect.grid(row=4,pady=20)
buttondisplay=Button(root,text="Display Attendance",command=get_name)
buttondisplay.grid(row=5,pady=20)
labeldisplayattendance=Label(root,text="",width=130,height=2)
labeldisplayattendance.grid(row=6,pady=20)

label=Label(root,text="NAME                         TIME",bg="white")
label.grid(row=7,pady=20)
labeldisplay=Label(root,text="",bg="white")
labeldisplay.grid(row=8)

root.mainloop()
